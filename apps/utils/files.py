# coding=utf8

"""
used for file process
"""

import os, logging
from zipfile import BadZipFile
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet import Worksheet

import settings
from openpyxl import Workbook, load_workbook

from utils import times, eggs
from utils.eggs import gen_uuid1
from utils.times import fn_timer

logger = logging.getLogger('django')


def is_file_exist(path):
    """ 判断文件或目录是否存在 """
    return os.path.exists(path)


def remove(path, fileName=None):
    """remove file from the filesystem"""
    if not fileName:
        fullpath = path
    else:
        fullpath = os.path.join(path, fileName)

    try:
        os.remove(fullpath)
        return True
    except OSError as e:
        logger.error("delete file %s error: %s" % (fullpath, e))
        return False


def save_file(file, base_dir, file_name):
    """保存文件
    @param file  传入的文件参数, request.FILES中获取的数据对象, file需要先经过rename处理, 以便获取到file.name
    :param base_dir:
    @param file_name  保存的目标文件名
     """

    if not file:
        return ''
    try:
        dest = open('%s/%s/%s' % (settings.MEDIA_ROOT, base_dir, file_name), 'wb+')
        for chunk in file.chunks():
            dest.write(chunk)
        dest.close()
    except Exception as e:
        logger.exception(e)
        dest.close()

    return file_name


def upload_file(file, base_dir, stored_name=None):
    """

    :param file: 文件
    :param base_dir: 文件在服务器的存储路径
    :param stored_name: 文件名
    :return: 上传的文件名和文件在服务器的存储路径组成字典{'name': 'ddd.txt', 'path': 'upload/documents/d434324324dsf23.txt'}
    """
    try:
        path = os.path.join(settings.MEDIA_ROOT, base_dir)
        if not os.path.exists(path):
            os.makedirs(path)
        # 对文件名重命名
        if not stored_name:
            stored_name = '%s%s' % (gen_uuid1(), os.path.splitext(file.name)[1])
        file_path = '%s%s' % (path, stored_name)
        destination = open(file_path, 'wb+')
        for chunk in file.chunks():
            destination.write(chunk)
        destination.close()
    except Exception as e:
        logger.info(e)
        return None
    data = {'name': file.name, 'path': base_dir + stored_name}
    return data


def single_upload_file(file, base_dir, stored_file_name):
    """

    :param file:
    :param base_dir:
    :param stored_file_name:
    :return: 由上传的文件名和
    """
    try:
        path = os.path.join(settings.MEDIA_ROOT, base_dir)
        if not os.path.exists(path):
            os.makedirs(path)
        file_path = '%s%s' % (path, stored_file_name)
        destination = open(file_path, 'wb+')

        for chunk in file.chunks():
            destination.write(chunk)
        destination.close()
    except Exception as e:
        logger.info(e)
        return None

    # 文件名称对应保存路径
    file_name_url_data = {
        file.name: '%s%s' % (base_dir, stored_file_name)
    }
    return file_name_url_data


def single_upload_file_test(file, base_dir, stored_file_name):
    """

    :param file:
    :param base_dir:
    :param storge_file_name:
    :return: 由上传的文件名和
    """
    try:
        path = os.path.join(settings.MEDIA_ROOT, base_dir)
        if not os.path.exists(path):
            os.makedirs(path)
        file_path = '%s%s' % (path, stored_file_name)
        destination = open(file_path, 'wb+')

        for chunk in file.chunks():
            destination.write(chunk)
        destination.close()
    except Exception as e:
        logger.info(e)
        return False
    return True


# def batch_update_files(file_dict, base_dir):
#
#     if not file_dict:
#         return None
#     path = None
#     try:
#         path = os.path.join(settings.MEDIA_ROOT, base_dir)
#         if not os.path.exists(path):
#             os.makedirs(path)
#     except Exception as e:
#         logger.info(e)
#         return None
#     try:
#         for file, stored_file_name in file_dict.items():
#             file_path = '%s/%s' % (path, stored_file_name)
#             destination = open(file_path, 'wb+')
#             for chunk in file.chunks():
#                 destination.write(chunk)
#             destination.close()
#     except Exception as e:
#         logger.info(e)
#         return None
#
#     return file.name, '%s%s' % (base_dir, stored_file_name)


def get_file_extension(path):
    """
    获取文件扩展名
    :param path:文件名或文件路径
    :return: 返回文件扩展名
    """
    return os.path.splitext(path)[1]


class ExcelBasedOXL(object):
    """
    基于Openpyxl库的excel封装
    """

    @staticmethod
    def open_excel(path):
        """
        打开excel
        :param path: the path to open or a file-like object
        :type path: string or a file-like object open in binary mode c.f., :class:`zipfile.ZipFile`
        :return: 返回Workbook对象，即excel文件对象
        """
        if not path:
            return False, 'path参数不能为空'
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            return True, workbook
        except (InvalidFileException, BadZipFile) as iex:
            logger.exception(iex)
            return False, '文件格式异常或文件损坏'
        except Exception as e:
            logger.exception(e)
            return False, '文件读取异常'

    @staticmethod
    def get_sheet(workbook, sheet_name):
        """
        获取Worksheet对象
        :return: 返回Worksheet对象
        """
        return workbook[sheet_name]

    @staticmethod
    def get_rows_len(sheet):
        """
        获取sheet的最大行数
        :param sheet: Worksheet对象
        :return: 返回行数
        """
        if not sheet:
            return None
        return sheet.max_row

    @staticmethod
    def get_cell_value(sheet, row, col):
        """
        获取单元格中内容
        :param sheet: Worksheet对象
        :param row: 行序列号，默认1是首行
        :param col: 列序列号，默认1是首列
        :return:
        """
        if not sheet:
            return None
        return sheet.cell(row, col).value

    @staticmethod
    def close(workbook):
        """
        关闭Workbook对象
        :return:
        """
        if workbook:
            workbook.close()

    @staticmethod
    def read_excel(wb, header_dict=None):
        """
        读取excel文件数据，带表头
        :param wb: Workbook对象
        :param header_dict: 表头字典，K:属性名；V:表头单元格数据
        :return:
            返回由读取结果（True/False）和读取的数据列表(excel_data)组成的Cuple
            excel_data是由多个sheet数据组成的List；
            sheet是由多个行数据组成的List；
            每行数据则是以header_dict的k作为键，对应单元格数据作为值组成的Dict

        """
        excel_data = []
        try:
            for sheet_name in wb.sheetnames:
                sheet_data = ExcelBasedOXL.read_sheet(wb[sheet_name], header_dict)
                if sheet_data:
                    excel_data.append(sheet_data)
            return True, excel_data
        except Exception as e:
            logger.exception(e)
            return False, "表头数据和指定的标准不一致或excel解析错误"

    @staticmethod
    def read_raw_excel(wb, header_dict_list=None):
        """
        读取excel文件数据，带表头
        :param wb: Workbook对象
        :param header_dict: 表头字典，K:属性名；V:表头单元格数据
        :return:
            返回由读取结果（True/False）和读取的数据列表(excel_data)组成的Cuple
            excel_data是由多个sheet数据组成的List；
            sheet是由多个行数据组成的List；
            每行数据则是以header_dict的k作为键，对应单元格数据作为值组成的Dict

        """
        sheet_header_keys_list = []
        sheet_data_list = []
        try:
            for index, sheet_name in enumerate(wb.sheetnames):
                header_dict = []
                if index < len(header_dict_list):
                    header_dict = header_dict_list[index]
                header_key_list = ExcelBasedOXL.read_sheet_header(wb[sheet_name],
                                                                  header_dict)
                sheet_header_keys_list.append(header_key_list)
        except Exception as e:
            logger.exception(e)
            return False, "表头数据和指定的标准不一致或Excel文件解析错误"
        try:
            for index, sheet_name in enumerate(wb.sheetnames):
                sheet_data = ExcelBasedOXL.read_raw_sheet(wb[sheet_name])
                if sheet_header_keys_list[index]:
                    sheet_data[0] = sheet_header_keys_list[index]
                sheet_data_list.append(sheet_data)
            return True, sheet_data_list
        except Exception as e:
            logger.exception(e)
            return False, "Excel文件解析异常"

    @staticmethod
    def read_sheet_header(ws: Worksheet, header_dict=None):
        """
        读取单个sheet数据，带表头
        :param ws: Worksheet对象
        :param header_dict: 表头字典，K:键，一般为model属性；V:对应表头单元格数据
        :return: 返回一个List对象，该List对象由一组Dictionary对象构成
        """
        if not ws or not header_dict:
            return []
        # 读取首行数据
        header_row_data = []
        for col in range(1, ws.max_column + 1):
            header_row_data.append(ws.cell(1, col).value)

        header_key_list = []  # 表头顺序对应的关键字列表
        # 判断首行数据是否和指定的标准一致，如果一致，按顺序封装表头关键字；否则抛出异常
        for header_cell_data in header_row_data:
            for item in header_dict.items():
                if item[1] == header_cell_data:
                    header_key_list.append(item[0])
                    break
        if len(header_key_list) != len(header_dict):
            # 抛出“表单的表头数据和指定的标准不一致，请检查”异常
            raise Exception('ExcelHeaderNotMatched')
        return header_key_list

    @staticmethod
    def read_raw_sheet(ws: Worksheet):
        """
        读取单个sheet数据, 不做任何封装处理
        :param ws:  Worksheet对象
        :return:
        """
        sheet_data = []
        if not ws:
            return sheet_data
        for row in ws.rows:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            sheet_data.append(row_data)
        return sheet_data

    @staticmethod
    @fn_timer
    def read_sheet(ws: Worksheet, header_dict=None):
        """
        读取单个sheet数据，带表头并将行数据封装为键值对
        :param ws: Worksheet对象
        :param header_dict: 表头字典，K:键，一般为model属性；V:对应表头单元格数据
        :return: 返回一个List对象，该List对象由一组Dictionary对象构成
        """
        if not ws:
            return []

        sheet_data = []

        # 读取不带表头的sheet
        if not header_dict:
            for row in ws.rows:
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                sheet_data.append(row_data)
            return sheet_data

        # 读取带表头的sheet
        # 读取首行数据
        header_row_data = []

        for row in ws.rows:
            for cell in row:
                header_row_data.append(cell.value)
            break

        header_keys = []  # 表头顺序对应的关键字列表
        # 判断首行数据是否和指定的标准一致，如果一致，按顺序封装表头关键字；否则抛出异常
        for header_cell_data in header_row_data:
            for item in header_dict.items():
                if item[1] == header_cell_data:
                    header_keys.append(item[0])
                    break

        if len(header_keys) != len(header_dict):
            # 抛出“表单的表头数据和指定的标准不一致，请检查”异常
            raise Exception('ExcelHeaderNotMatched')

        # 读取业务数据
        for row_index, row in enumerate(ws.rows):
            row_data = {}
            if row_index == 0:
                continue
            for cell_index, cell in enumerate(row):
                key = header_keys[cell_index]
                if cell.value and cell.is_date:
                    value = cell.value.strftime('%Y-%m-%d')
                else:
                    value = cell.value if cell.value else ''
                row_data[key] = value
            sheet_data.append(row_data)

        return sheet_data

    @staticmethod
    def gen_sheet_name(param):
        return param.__unicode__()

    @staticmethod
    def gen_workbook():
        return Workbook()

    @staticmethod
    def export_excel(base_dir, file_name, records_list, sheet_names=[], header_rows=[], ):
        """
        导出Excel文件
        :param base_dir: 文件存放基本路径
        :param file_name: 文件名（不带后缀）
        :param records_list: sheet数据list， 列表的每个元素，代表一个sheet的数据，每个元素也是list对象，该列表存放每行的数据
        :param sheet_names:  自定义的sheet名称列表，sheet_names[index] 和records_list[index]相对应
        :param header_rows:  标题行列表, header_rows[index] 和records_list[index]相对应
        :return:
        """
        wb = Workbook()
        postfix = 'xlsx'
        filename = gen_filename(file_name, postfix, date_format=None)
        path = os.path.join(settings.MEDIA_ROOT, base_dir)
        if not os.path.exists(path):
            os.makedirs(path)
        file_path = '%s%s' % (path, filename)
        wb.remove(wb.active)
        try:
            for index, records in enumerate(records_list):
                ws = wb.create_sheet(('Sheet' + str(index + 1)), index)
                if sheet_names:
                    if index <= len(sheet_names) - 1:
                        if sheet_names[index]:
                            ws.title = sheet_names[index]
                if header_rows and index <= len(header_rows) - 1:
                    ExcelBasedOXL.write_to_sheet_by_row(ws, records, header_rows[index])
                else:
                    ExcelBasedOXL.write_to_sheet_by_row(ws, records)
            wb.save(filename=file_path)
            return '%s.%s' % (file_name, postfix), '%s%s' % (base_dir, filename)
        except Exception as e:
            logger.exception(e)
            raise Exception("Write data to file exception!")
        finally:
            wb.close()

    @staticmethod
    def write_to_sheet_by_row(worksheet, data_rows, header_row=None):
        """
        添加数据到Worksheet对象中，按行添加
        :param worksheet: Worksheet对象
        :param header_row: 表头列表
        :param data_rows: data_rows为行数据列表，每个元素是一行（row）数据，也是list对象
        :return:
        """
        try:
            if header_row:
                worksheet.append(header_row)
            if data_rows:
                for data in data_rows:
                    worksheet.append(data)
        except Exception as e:
            logger.exception(e)
            raise Exception("Write data to sheet exception!")


def gen_filename(file_name, postfix, date_format='%Y%m%d%H%M%S%f', uuid_first=False):
    """
    生成文件名
    :param file_name: 文件名称（不带后缀）string
    :param postfix: 文件后缀 string
    :param date_format: 时间戳转换格式， string
    :return:
    """
    if not file_name or not postfix:
        return None
    if uuid_first:
        return '%{filename}_{uuid}.{postfix}'.format(
            filename=file_name, uuid=eggs.gen_uuid1(), postfix=postfix
        )
    if not date_format:
        return '%{filename}.{postfix}'.format(filename=file_name, postfix=postfix)
    return '%{filename}_{timestamp}.{postfix}'.format(
        filename=file_name, timestamp=times.datetime_to_str(times.now(), date_format),
        postfix=postfix
    )


def file_read_iterator(file_name, chunk_size=1024):
    """文件读取迭代器"""
    try:
        with open(file_name, 'rb') as f:
            while True:
                c = f.read(chunk_size)
                if c:
                    yield c
                else:
                    break
    except Exception as e:
        print('Read file exception!')
        raise Exception('Read file exception!')











